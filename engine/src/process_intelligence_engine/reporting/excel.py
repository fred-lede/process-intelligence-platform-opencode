"""Excel report generator."""
from __future__ import annotations
from datetime import datetime
from typing import Any

from .base import ReportGenerator
from .models import ReportData


class ExcelReportGenerator(ReportGenerator):
    """Generate Excel report."""
    
    def generate(self) -> bytes:
        """Generate Excel report as bytes."""
        try:
            from io import BytesIO
            from openpyxl import Workbook
        except ImportError:
            raise ImportError("openpyxl is required for Excel generation. Install with: pip install openpyxl")
        
        output = BytesIO()
        wb = Workbook()
        
        # Sheet 1: Project Info
        ws1 = wb.active
        ws1.title = "專案資訊"
        ws1.append(["專案名稱", self.data.project_name])
        ws1.append(["操作者", self.data.operator])
        ws1.append(["產生時間", self.data.created_at.strftime("%Y-%m-%d %H:%M:%S")])
        ws1.append(["資料集 ID", self.data.dataset_id])
        ws1.append(["來源檔案", self.data.source_file])
        ws1.append(["資料列數", self.data.row_count])
        ws1.append(["欄位數", self.data.column_count])
        
        # Sheet 2: Field Roles
        ws2 = wb.create_sheet("欄位角色")
        ws2.append(["欄位名稱", "角色", "信心度", "資料型態"])
        for field in self.data.fields:
            ws2.append([
                field.get("name", ""),
                field.get("role", ""),
                field.get("confidence", 0),
                field.get("data_type", ""),
            ])
        
        # Sheet 3: Model Comparison
        if self.data.model_comparison:
            ws3 = wb.create_sheet("模型比較")
            ws3.append(["模型 ID", "模型類型", "R²", "RMSE", "MAE", "Adj R²", "狀態"])
            for model in self.data.model_comparison:
                ws3.append([
                    model.get("model_id", ""),
                    model.get("model_type", ""),
                    model.get("metrics", {}).get("r2", ""),
                    model.get("metrics", {}).get("rmse", ""),
                    model.get("metrics", {}).get("mae", ""),
                    model.get("metrics", {}).get("adj_r2", ""),
                    model.get("status", ""),
                ])
        
        # Sheet 4: Interactions
        if self.data.interactions.get("matrix"):
            ws4 = wb.create_sheet("交互作用")
            factors = self.data.interactions.get("factors", [])
            ws4.append([""] + factors)
            matrix = self.data.interactions.get("matrix", [])
            for i, factor in enumerate(factors):
                row = [factor]
                for j, val in enumerate(matrix[i]):
                    row.append(val)
                ws4.append(row)
        
        # Sheet 5: Recommendations
        if self.data.recommendations:
            ws5 = wb.create_sheet("實驗建議")
            ws5.append(["類型", "優先級", "因子", "說明"])
            for rec in self.data.recommendations:
                ws5.append([
                    rec.get("type", ""),
                    rec.get("priority", ""),
                    ", ".join(rec.get("factors", [])),
                    rec.get("reason", ""),
                ])
        
        wb.save(output)
        output.seek(0)
        return output.read()
